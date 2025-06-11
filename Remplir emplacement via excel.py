# -*- coding: utf-8 -*-
"""
Created on Wed Mar 19 20:39:14 2025

@author: User
"""

import mysql.connector
from openpyxl import load_workbook

# Paramètres de connexion à la base MySQL
config = {
    'user': 'root',
    'password': 'VoltR99!',
    'host': 'localhost',
    'database': 'bdd_voltr_cloud',
    'auth_plugin': 'mysql_native_password',
    'raise_on_warnings': True
}

# 1) Lire les numéros de série depuis le fichier Excel
excel_file_path = r'C:\Users\User\Desktop\P_test.xlsx'

# On suppose qu’il n’y a qu’une seule feuille ou qu’on prend la 1ère feuille active
workbook = load_workbook(excel_file_path)
sheet = workbook.active

# Récupérer les numéros de série dans une liste
# On saute la première ligne si c’est un header
numero_series = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    # Supposons que le numéro de série soit en colonne A (donc row[0])
    if row[0] is not None:
        numero_series.append(str(row[0]).strip())

try:
    # 2) Connexion à la base MySQL
    cnx = mysql.connector.connect(**config)
    cursor = cnx.cursor()

    # 3) Récupérer les ID des emplacements libres, classés par ordre
    select_query = """
        SELECT id
        FROM emplacements
        WHERE est_occupe = 0
        ORDER BY id
    """
    cursor.execute(select_query)
    free_spots = [row[0] for row in cursor.fetchall()]

    # 4) Pour chaque emplacement libre, assigner le prochain numéro de série
    update_query = """
        UPDATE emplacements
        SET est_occupe = 1,
            numero_serie = %s,
            date_attribution = NOW()
        WHERE id = %s
    """
    count_updated = 0
    for i, spot_id in enumerate(free_spots):
        if i >= len(numero_series):
            break  # Plus de numéro de série à affecter
        cursor.execute(update_query, (numero_series[i], spot_id))
        count_updated += 1

    cnx.commit()
    print(f"{count_updated} emplacements ont été mis à jour.")

except mysql.connector.Error as err:
    print(f"Erreur MySQL : {err}")

finally:
    if cursor:
        cursor.close()
    if cnx:
        cnx.close()
